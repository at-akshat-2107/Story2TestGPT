import streamlit as st
import json
import re
import pandas as pd
import os
import random
import string
from groq import Groq
import nltk
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
import textstat
from rouge_score import rouge_scorer
from tabulate import tabulate
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import shutil
from datetime import datetime
from zipfile import ZipFile
from PIL import Image
from collections import Counter
import shap
import lime.lime_text
import numpy as np
from fpdf import FPDF
import tempfile

saved_data = ""
# Initialize global variables and download necessary NLTK resources
MODEL_NAME = "llama3-70b-8192"
TECHNIQUE_NAME = "Chain of Thought Prompting"
nltk.download('punkt')
nltk.download('stopwords')
os.environ['GROQ_API_KEY'] = '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'
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

# Define a PDF class for generating a document with a cover and test cases
class PDF(FPDF):
    def header(self):
        if self.page_no() > 1:
            self.set_font('Times', 'B', 14)
            self.cell(0, 10, 'Story2TestGPT Generated Test Cases', 0, 1, 'C')
            self.ln(2)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(5)

    def footer(self):
        if self.page_no() > 1:
            self.set_y(-15)
            self.set_font('Times', 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def add_cover_page(self):
        self.add_page()
        self.set_fill_color(230, 230, 230)
        self.rect(0, 0, 210, 297, 'F')
        self.set_font('Times', 'B', 28)
        self.set_text_color(0, 51, 102)
        self.cell(0, 100, 'Story2TestGPT Generated Test Cases', 0, 1, 'C')
        self.set_font('Times', 'B', 16)
        self.cell(0, 10, 'Creator of Story2TestGPT - Akshat Mehta', 0, 1, 'C')
        self.set_font('Times', '', 14)
        self.cell(0, 10, 'AI/ML Engineer at Inexture Solutions Limited', 0, 1, 'C')
        self.ln(20)
        self.set_line_width(0.5)
        self.set_draw_color(169, 169, 169)
        self.line(20, self.get_y(), 190, self.get_y())
        self.ln(10)
        self.set_font('Times', '', 12)
        self.multi_cell(0, 10, 'This document contains test cases generated using the Story2TestGPT system...', align='C')
        self.ln(5)
        self.line(20, self.get_y(), 190, self.get_y())
        self.ln(5)
        self.multi_cell(0, 10, 'Disclaimer: The generated test cases are based on the input data...', align='C')
        self.ln(5)
        self.line(20, self.get_y(), 190, self.get_y())
        self.ln(5)
        self.multi_cell(0, 10, 'For more information, contact: Akshat Mehta at at.akshat.mehta.2107@gmail.com', align='C')

def save_pdf(content):
    pdf = PDF()
    pdf.set_left_margin(10)
    pdf.set_right_margin(10)
    pdf.add_cover_page()
    pdf.add_page()
    paragraphs = content.split("<div class='divider'></div>")
    for paragraph in paragraphs:
        paragraph = paragraph.replace("<br>", "\n")
        paragraph = re.sub('<[^<]+?>', '', paragraph).strip()
        if paragraph:
            paragraph = paragraph.encode('latin1', 'ignore').decode('latin1')
            lines = paragraph.split("\n")
            for line in lines:
                if ':' in line:
                    label, value = line.split(":", 1)
                    pdf.set_font('Times', 'B', 12)
                    pdf.cell(40, 10, label.strip() + ":", ln=0)
                    pdf.set_font('Times', '', 12)
                    pdf.multi_cell(0, 10, value.strip(), align='L')
                else:
                    pdf.set_font('Times', '', 12)
                    pdf.multi_cell(0, 10, line, align='L')
            pdf.ln()
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf.output(temp_file.name, 'F')
    return temp_file.name

def generate_random_id(length=6):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

def preprocess_data_for_streamlit(project_name, project_description, feature_name, feature_description, userstory_title, acceptance_criteria):
    preprocessed_data = {
        "project_name": project_name,
        "project_description": project_description,
        "feature_name": feature_name,
        "feature_description": feature_description,
        "userstory_title": userstory_title,
        "acceptance_criteria": acceptance_criteria
    }
    return preprocessed_data

def create_prompt(preprocessed_data):
    prompt = f"""
    Persona: Detailed-Oriented QA Engineer
    As a QA Engineer, I am meticulous, ensuring every acceptance criterion is thoroughly covered in test cases. I approach test case creation systematically, emphasizing clarity, completeness, and alignment with user stories. My goal is to generate comprehensive, detailed test cases that guarantee software functionality meets user expectations.
    You are functioning as a Quality Assurance Engineer tasked with generating detailed test cases for a software project based on the following input data. Use the provided Test Case Creation Template to structure your test cases effectively. Generate at least 5 test cases, ensuring each acceptance criterion from the user story is covered by at least one test case.
    Think through the process of generating test cases step by step:
    1. Carefully read and understand the project details and the user story provided in the input data.
    2. Identify and list all the acceptance criteria mentioned in the user story.
    3. For each acceptance criterion, think about what functionalities need to be tested to ensure that the criterion is met.
    4. Create test cases that cover these functionalities. Each test case should be unique and should include:
       - A unique Test Case ID
       - A clear and concise Test Case Title that begins with "Verify that..."
       - A detailed Test Case Description that explains the purpose and scope of the test
       - The appropriate Test Suite or module to which the test case belongs
       - The Priority level of the test case (High, Medium, Low)
       - Any Preconditions that must be met before executing the test case
       - The Test Data required for the test case, if any
       - The Test Steps to be followed to execute the test case, in a clear and logical sequence
       - The Postconditions that should be true after the test case is executed
       - The Expected Result of the test case, which should be a clear and measurable outcome
       - The Severity of the test case (Blocker, Critical, Major, Minor, Trivial)
       - The Type of Testing (Functional, Smoke, Regression, Load, Performance)
       - The Test Case Behaviour (Positive, Negative)
    5. Ensure that the test cases are comprehensive and cover all aspects of the acceptance criteria effectively.
    6. Review the generated test cases to check for completeness and accuracy.
    Input Data:
    Project Name: {preprocessed_data['project_name']}
    Project Description: {preprocessed_data['project_description']}
    Feature Name: {preprocessed_data['feature_name']}
    Feature Description: {preprocessed_data['feature_description']}
    User Story Title: {preprocessed_data['userstory_title']}
    Acceptance Criteria: {", ".join(preprocessed_data['acceptance_criteria'])}
    Test Case Creation Template:
    Test Case ID: (Provide a unique identifier for each test case)
    Test Case Title: (Must start with "Verify that...", must be between 20-30 words and similar to the acceptance criteria which you covered)
    Test Case Description: (A brief description of the test case, minimum in between 70-100 words)
    Test Suite: (Name of the test suite or module)
    Test Priority: (Priority level: High, Medium, Low)
    Preconditions: (List any prerequisites before executing the test case, min 3, max 5 items, Format in Ordered list)
    Test Data: (Specify data required for execution; if none, state "No test data needed")
    Test Steps: (List steps for executing the test case, min 5 to max 6, Format in Ordered list of all steps)
    Postconditions: (Any follow-up actions after test execution, min 3, max 5 items, Format in Ordered list)
    Expected Result: (Overall expected outcome of the test case)
    Severity: (Blocker, Critical, Major, Minor, Trivial)
    Priority: (High, Medium, Low - Try to Cover all Types)
    Type of Testing: (Functional, Smoke, Regression, Load, Performance - Try to Cover all Types)
    Test Case Behaviour: (Positive, Negative - Try to Cover all Types)
    Generate the test cases, ensuring thoroughness and coverage according to the acceptance criteria given.
    """
    return prompt

def generate_test_cases_with_groq(input_text):
    chat_completion = client.chat.completions.create(
        messages=[{"role": "user", "content": input_text}], model=MODEL_NAME)
    generated_test_cases = chat_completion.choices[0].message.content
    return generated_test_cases

def generate_explanations_lime(generated_test_cases, reference_test_cases):
    explainer = lime.lime_text.LimeTextExplainer(class_names=["Not Matching", "Matching"])
    explanations = []

    def prediction_fn(texts):
        return np.array([[1 if text == reference else 0, 0 if text == reference else 1] for text in texts])

    for generated, reference in zip(generated_test_cases, reference_test_cases):
        exp = explainer.explain_instance(generated, prediction_fn, num_features=10, labels=(1,))
        explanation_text = "LIME Explanation: " + " ".join([f"{feature}" for feature, weight in exp.as_list(label=1)])
        explanations.append(explanation_text)
    return explanations

def generate_explanations_shap(generated_test_cases, reference_test_cases):
    def prediction_fn(texts):
        return np.array([1 if text in reference_test_cases else 0 for text in texts])

    explainer = shap.Explainer(prediction_fn, shap.maskers.Text())
    explanations = []

    for generated in generated_test_cases:
        shap_values = explainer([generated])
        explanation_text = "SHAP Explanation: " + " ".join([f"{feature}" for feature in generated.split()])
        explanations.append(explanation_text)

    return explanations

def generate_xai_explanations(generated_test_cases, reference_test_cases):
    lime_explanations = generate_explanations_lime(generated_test_cases, reference_test_cases)
    shap_explanations = generate_explanations_shap(generated_test_cases, reference_test_cases)
    
    combined_explanations = []
    for lime, shap in zip(lime_explanations, shap_explanations):
        combined_explanations.append(f"{lime}\n{shap}")
    
    return combined_explanations

def summarize_explanations_with_llm(explanation):
    prompt = f"""
    Persona: A meticulous and experienced QA Engineer. Your task is to craft a professional and concise summary that highlights the critical importance and reasoning behind the generated test case. The summary should use standard, easy-to-understand language, directly providing the reason and significance of the test case in 80-100 words.
    Structure the summary as follows in one paragraph:
    - Begin by identifying the core focus of the test case.
    - Explain why this test case is crucial for ensuring the quality and functionality of the feature.
    - Highlight any specific conditions or aspects that contribute to its importance.
    Please ensure the summary directly addresses the purpose, reason, and critical importance of the test case based on the following explanation:
    {explanation}
    """
    chat_completion = client.chat.completions.create(
        messages=[{"role": "user", "content": prompt}], model=MODEL_NAME)
    summary = chat_completion.choices[0].message.content.strip()

    summary_lines = summary.splitlines()
    summary = " ".join(line.strip() for line in summary_lines if len(line.strip()) > 0 and not line.lower().startswith(('here is', 'this is', 'summary', 'the following')))
    formatted_summary = summary.capitalize()
    
    formatted_summary = formatted_summary.strip()
    formatted_summary = ' '.join(formatted_summary.split())
    if not formatted_summary.endswith('.'):
        formatted_summary += '.'
    formatted_summary = re.sub(r'\s+', ' ', formatted_summary)
    return formatted_summary

def format_generated_test_cases(preprocessed_data, generated_test_cases, reference_test_cases):
    custom_css = """
    <style>
.custom-container {
    border: 3px solid #333; /* Dark border */
    border-radius: 10px; /* Rounded corners */
    background-color: #f9f9f9; /* Light background color */
    padding: 20px; /* Padding inside the container */
    margin: 20px 0; /* Margin outside the container */
    font-family: 'Times New Roman', Times, serif;
}
.custom-title {
    font-size: 16px;
    font-weight: bold;
    margin-bottom: 10px;
}
.custom-label {
    font-size: 14px;
    font-weight: bold;
    display: inline-block;
    margin-top: 10px;
}
.custom-value {
    font-size: 14px;
    font-weight: normal;
    display: inline-block;
    margin-left: 10px;
}
.divider {
    border-top: 2px solid #bbb;
    margin-top: 10px;
    margin-bottom: 10px;
}
.scrollable-output {
    max-height: 90vh;
    overflow-y: scroll;
    padding-right: 15px;
}
</style>
    """
    st.markdown(custom_css, unsafe_allow_html=True)
    
    formatted_output = f"<div class='custom-container'>"
    formatted_output += f"<div class='custom-title'>As per your request...For project {preprocessed_data['project_name']}'s Feature {preprocessed_data['feature_name']} 5 Test Cases are Generated, Kindly Check!</div>"
    formatted_output += "<div class='divider'></div>"
    formatted_output += "<div>"
    formatted_output += f"<span class='custom-label'>Project Name: </span><span class='custom-value'>{preprocessed_data['project_name']}</span><br>"
    formatted_output += f"<span class='custom-label'>Project Description: </span><span class='custom-value'>{preprocessed_data['project_description']}</span><br>"
    formatted_output += f"<span class='custom-label'>Feature Name: </span><span class='custom-value'>{preprocessed_data['feature_name']}</span><br>"
    formatted_output += f"<span class='custom-label'>Feature Description: </span><span class='custom-value'>{preprocessed_data['feature_description']}</span><br>"
    formatted_output += f"<span class='custom-label'>User Story Title: </span><span class='custom-value'>{preprocessed_data['userstory_title']}</span><br>"
    formatted_output += "<span class='custom-label'>Acceptance Criteria:</span><br>"
    
    unique_criteria = list(set(preprocessed_data['acceptance_criteria']))
    for i, criterion in enumerate(unique_criteria, 1):
        formatted_output += f"<span class='custom-value'>{i}. {criterion}</span><br>"
    
    formatted_output += "</div>"
    test_cases_split = generated_test_cases.split("Test Case ID:")
    test_case_count = 0

    for test_case in test_cases_split[1:]:
        test_case_count += 1
        test_case_data = {}
        test_case_data['id'] = generate_random_id()
        title_match = re.search(r"Test Case Title: (.*)", test_case)
        test_case_data['title'] = title_match.group(1).strip() if title_match else "N/A"
        description_match = re.search(r"Test Case Description: (.*)", test_case)
        test_case_data['description'] = description_match.group(1).strip() if description_match else "N/A"
        suite_match = re.search(r"Test Suite: (.*)", test_case)
        test_case_data['suite'] = suite_match.group(1).strip() if suite_match else "N/A"
        priority_match = re.search(r"Test Priority: (.*)", test_case)
        test_case_data['priority'] = priority_match.group(1).strip() if priority_match else "N/A"
        preconditions_match = re.search(r"Preconditions:(.*)Test Data:", test_case, re.DOTALL)
        if preconditions_match:
            test_case_data['preconditions'] = preconditions_match.group(1).strip().split('\n')
        else:
            test_case_data['preconditions'] = ["N/A"]
        test_data_match = re.search(r"Test Data: (.*)", test_case)
        test_case_data['test_data'] = test_data_match.group(1).strip() if test_data_match else "N/A"
        test_steps_match = re.search(r"Test Steps:(.*)Postconditions:", test_case, re.DOTALL)
        if test_steps_match:
            test_case_data['test_steps'] = test_steps_match.group(1).strip().split('\n')
        else:
            test_case_data['test_steps'] = ["N/A"]
        postconditions_match = re.search(r"Postconditions:(.*)Expected Result:", test_case, re.DOTALL)
        if postconditions_match:
            test_case_data['postconditions'] = postconditions_match.group(1).strip().split('\n')
        else:
            test_case_data['postconditions'] = ["N/A"]
        expected_result_match = re.search(r"Expected Result: (.*)", test_case)
        test_case_data['expected_result'] = expected_result_match.group(1).strip() if expected_result_match else "N/A"
        severity_match = re.search(r"Severity: (.*)", test_case)
        test_case_data['severity'] = severity_match.group(1).strip() if severity_match else "N/A"
        type_of_testing_match = re.search(r"Type of Testing: (.*)", test_case)
        test_case_data['type_of_testing'] = type_of_testing_match.group(1).strip() if type_of_testing_match else "N/A"
        behaviour_match = re.search(r"Test Case Behaviour: (.*)", test_case)
        test_case_data['behaviour'] = behaviour_match.group(1).strip() if behaviour_match else "N/A"
        
        formatted_output += "<div class='divider'></div>"
        formatted_output += f"<div class='custom-label'>Test Case No:</div><div class='custom-value'>{test_case_count}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Case Title:</div><div class='custom-value'>{test_case_data['title']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Case ID:</div><div class='custom-value'>{test_case_data['id']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Case Description:</div><div class='custom-value'>{test_case_data['description']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Suite:</div><div class='custom-value'>{test_case_data['suite']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Priority:</div><div class='custom-value'>{test_case_data['priority']}</div><br>"
        formatted_output += f"<div class='custom-label'>Preconditions:</div><br>"
        for i, precondition in enumerate(test_case_data['preconditions'], 1):
            formatted_output += f"<div class='custom-value'>{precondition.strip()}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Data:</div><div class='custom-value'>{test_case_data['test_data']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Steps:</div><br>"
        for i, step in enumerate(test_case_data['test_steps'], 1):
            formatted_output += f"<div class='custom-value'>{step.strip()}</div><br>"
        formatted_output += f"<div class='custom-label'>Postconditions:</div><br>"
        for i, postcondition in enumerate(test_case_data['postconditions'], 1):
            formatted_output += f"<div class='custom-value'>{postcondition.strip()}</div><br>"
        formatted_output += f"<div class='custom-label'>Expected Result:</div><div class='custom-value'>{test_case_data['expected_result']}</div><br>"
        formatted_output += f"<div class='custom-label'>Severity:</div><div class='custom-value'>{test_case_data['severity']}</div><br>"
        formatted_output += f"<div class='custom-label'>Type of Testing:</div><div class='custom-value'>{test_case_data['type_of_testing']}</div><br>"
        formatted_output += f"<div class='custom-label'>Test Case Behaviour:</div><div class='custom-value'>{test_case_data['behaviour']}</div><br>"
        
        xai_explanation = generate_xai_explanations([test_case], reference_test_cases)
        summarized_explanation = summarize_explanations_with_llm(xai_explanation[0])
        formatted_output += f"<div class='custom-label'>Explanation:</div><div class='custom-value'>{summarized_explanation}</div><br>"
    
    formatted_output += "</div>"  # Close the custom container div
    formatted_output = f"<div class='scrollable-output'>{formatted_output}</div>"
    st.session_state.pdf_path = save_pdf(formatted_output)
    return formatted_output

st.set_page_config(layout="wide")
background_image = "img/1.jpg"

page_bg_img = f"""
<style>
[data-testid="stAppViewContainer"] > .main {{
background-image: url("{background_image}");
background-size: cover;
background-position: top left;
background-repeat: no-repeat;
background-attachment: fixed;
}}
[data-testid="stHeader"] {{
background: rgba(0, 0, 0, 0);
}}
.st-bb {{
    background-color: rgba(255, 255, 255, 0.8);
}}
.css-145kmo2 {{
    background-color: rgba(255, 255, 255, 0.8);
}}
</style>
"""

st.markdown(page_bg_img, unsafe_allow_html=True)

st.markdown("<h1 style='text-align: center;'>Story2TestGPT</h1>", unsafe_allow_html=True)

input_col, output_col = st.columns([1, 1.5])

with input_col:
    project_name = st.text_input("Project Name", "", help="Enter the name of the project")
    if not project_name:
        st.markdown("<span style='color:red;'>Project Name is required!</span>", unsafe_allow_html=True)
    elif len(project_name) < 2 or len(project_name) > 200:
        st.markdown("<span style='color:red;'>Project Name must be between 2 and 200 characters!</span>", unsafe_allow_html=True)

    project_description = st.text_area("Project Description", "", help="Enter a description for the project (50-10000 characters)")
    if not project_description:
        st.markdown("<span style='color:red;'>Project Description is required!</span>", unsafe_allow_html=True)
    elif len(project_description) < 2 or len(project_description) > 5000:
        st.markdown("<span style='color:red;'>Project Description must be between 2 and 5000 characters!</span>", unsafe_allow_html=True)

    feature_name = st.text_input("Feature Name", "", help="Enter the name of the feature")
    if not feature_name:
        st.markdown("<span style='color:red;'>Feature Name is required!</span>", unsafe_allow_html=True)
    elif len(feature_name) < 2 or len(feature_name) > 200:
        st.markdown("<span style='color:red;'>Feature Name must be between 2 and 200 characters!</span>", unsafe_allow_html=True)

    feature_description = st.text_area("Feature Description", "", help="Enter a description for the feature (50-10000 characters)")
    if not feature_description:
        st.markdown("<span style='color:red;'>Feature Description is required!</span>", unsafe_allow_html=True)
    elif len(feature_description) < 2 or len(feature_description) > 5000:
        st.markdown("<span style='color:red;'>Feature Description must be between 2 and 5000 characters!</span>", unsafe_allow_html=True)

    userstory_title = st.text_area("User Story Title", "", help="Enter the user story title in the format 'As a...I want...so that...'. Example: As a user, I want to log in so that I can access my account.")
    if not userstory_title:
        st.markdown("<span style='color:red;'>User Story Title is required!</span>", unsafe_allow_html=True)
    elif not re.match(r'(?i)(.*as a.*i want.*so that.*|.*as a.*so that.*i want.*|.*i want.*as a.*so that.*|.*i want.*so that.*as a.*|.*so that.*as a.*i want.*|.*so that.*i want.*as a.*)', userstory_title):
        st.markdown("<span style='color:red;'>User Story Title must be in the format 'As a...I want...so that...'</span>", unsafe_allow_html=True)

    acceptance_criteria = []
    for i in range(5):
        criterion = st.text_input(f"Acceptance Criterion {i+1}", "", help=f"Enter acceptance criterion {i+1}")
        if not criterion:
            st.markdown(f"<span style='color:red;'>Acceptance Criterion {i+1} is required!</span>", unsafe_allow_html=True)
        acceptance_criteria.append(criterion)

    # Initialize saved_data at the top of your Streamlit code
    
    if st.button("Generate Test Cases", key="generate", on_click=st.empty):
        if (project_name and project_description and feature_name and feature_description and userstory_title and all(acceptance_criteria) and re.match(r'(?i)(.*as a.*i want.*so that.*|.*as a.*so that.*i want.*|.*i want.*as a.*so that.*|.*i want.*so that.*as a.*|.*so that.*as a.*i want.*|.*so that.*i want.*as a.*)', userstory_title) and 2 <= len(project_name) <= 200 and 2 <= len(project_description) <= 5000 and 2 <= len(feature_name) <= 200 and 2 <= len(feature_description) <= 5000):
            with st.spinner('Generating Test Cases...'):
                preprocessed_data = preprocess_data_for_streamlit(
                    project_name, project_description, feature_name, feature_description, userstory_title, acceptance_criteria
                )
                input_text = create_prompt(preprocessed_data)
                generated_test_cases = generate_test_cases_with_groq(input_text)
                reference_test_cases = [test_case.strip() for test_case in acceptance_criteria if test_case.strip()]
                formatted_output = format_generated_test_cases(preprocessed_data, generated_test_cases, reference_test_cases)
                st.session_state['generated_test_cases'] = formatted_output
                saved_data = st.session_state['generated_test_cases'] 
        else:
            st.error("Please check the highlighted fields and correct any errors. Ensure all required information is provided, and that all inputs meet the specified length and format requirements before generating test cases!")

with output_col:
    # Calculate the number of filled fields
    total_fields = 10  # Update this with the total number of input fields if different
    total_filled = len([field for field in [project_name, project_description, feature_name, feature_description, userstory_title] + acceptance_criteria if field])

    # Calculate the progress percentage
    progress_percentage = (total_filled / total_fields) * 100

    # Display the progress bar with percentage information
    st.markdown(f"Input Data Completion Progress: **{int(progress_percentage)}%**")
    progress_bar = st.progress(0)
    progress_bar.progress(int(progress_percentage))
    

    if "generated_test_cases" in st.session_state:
        st.subheader("Generated Test Cases by Story2TestGPT")
        st.markdown(saved_data, unsafe_allow_html=True)
        
        # Store the clicked state in the session state
        if "downloaded" not in st.session_state:
            st.session_state.downloaded = False
        
        if "pdf_path" in st.session_state and st.session_state.pdf_path:
            st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
            with open(st.session_state.pdf_path, "rb") as pdf_file:
                if st.download_button(
                    label="Download Generated Test Cases",
                    data=pdf_file,
                    file_name="generated_test_cases.pdf",
                    mime="application/pdf"
                ):
                    st.session_state.downloaded = True  # Set to True when button is clicked

        # Display success message after download
        if st.session_state.downloaded:
            st.success("PDF downloaded successfully!")




